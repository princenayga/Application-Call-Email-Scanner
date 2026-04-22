"""
clean_listserv.py
-----------------
Reads the fellowship_replies.xlsx and the original listserv CSV.
Produces a 3-sheet Excel file:

  Sheet 1 - Cleaned Listserv  : safe to blast next cycle (bounces removed)
  Sheet 2 - Needs Review      : No Longer Works There + Declined -- decide manually
  Sheet 3 - Removed           : confirmed dead bounces, for your records
"""

import os
import sys
import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependency: openpyxl  ->  pip install openpyxl")
    sys.exit(1)

# ─── CONFIG ──────────────────────────────────────────────────────────────────
EXCEL_INPUT  = os.path.join("output", "fellowship_replies.xlsx")
LISTSERV_CSV = "Filtered Listserv for Feb 2026 Applications - Sheet1.csv"
OUTPUT_FILE  = os.path.join("output", "cleaned_listserv.xlsx")

# Only auto-remove confirmed dead addresses
AUTO_REMOVE_CATEGORIES = {
    "Bounce / Delivery Failure",
}

# Flag for manual review — do NOT auto-remove
REVIEW_CATEGORIES = {
    "No Longer Works There",   # may have forwarded a new contact
    "Declined / Not Interested",  # may be "post to board" or "resend directly" cases
}
# ─────────────────────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF")
BOUNCE_FILL  = PatternFill("solid", fgColor="FFC7CE")   # light red
REVIEW_FILL  = PatternFill("solid", fgColor="FFE699")   # amber


def load_listserv(csv_path):
    emails = []
    if not os.path.exists(csv_path):
        print(f"Listserv CSV not found: {csv_path}")
        sys.exit(1)
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        for line in f:
            addr = line.strip().lower()
            if addr and "@" in addr:
                emails.append(addr)
    return emails


def load_classified(excel_path):
    """
    Returns two dicts, both keyed by lowercase email address:
      removals : email -> category   (auto-remove bounces)
      reviews  : email -> (category, sender_name, summary, action)
    """
    if not os.path.exists(excel_path):
        print(f"Excel report not found: {excel_path}")
        print("Run fellowship_scanner.py first.")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path)
    if "Replies" not in wb.sheetnames:
        print("No 'Replies' sheet found in the Excel file.")
        sys.exit(1)

    ws = wb["Replies"]

    # Columns (0-based): #, Sender Name, Email Address, Institution,
    #                    Date Received, Gmail Category, Category, Summary, Action Needed
    c_name    = 1
    c_email   = 2
    c_cat     = 6
    c_summary = 7
    c_action  = 8

    removals = {}
    reviews  = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        email    = str(row[c_email]).strip().lower()   if row[c_email]   else ""
        name     = str(row[c_name]).strip()            if row[c_name]    else ""
        category = str(row[c_cat]).strip()             if row[c_cat]     else ""
        summary  = str(row[c_summary]).strip()         if row[c_summary] else ""
        action   = str(row[c_action]).strip()          if row[c_action]  else ""

        if not email:
            continue
        if category in AUTO_REMOVE_CATEGORIES:
            removals[email] = category
        elif category in REVIEW_CATEGORIES:
            reviews[email] = (category, name, summary, action)

    return removals, reviews


def style_header(ws, cols):
    for col_idx, header in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 22)
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def main():
    print(f"Loading listserv : {LISTSERV_CSV}")
    listserv = load_listserv(LISTSERV_CSV)
    print(f"  -> {len(listserv)} address(es)\n")

    print(f"Loading replies  : {EXCEL_INPUT}")
    removals, reviews = load_classified(EXCEL_INPUT)
    print(f"  -> {len(removals)} confirmed bounce(s) to auto-remove")
    print(f"  -> {len(reviews)} address(es) needing manual review\n")

    kept         = [e for e in listserv if e not in removals and e not in reviews]
    review_rows  = [(e,) + reviews[e] for e in listserv if e in reviews]
    removed_rows = [(e, removals[e]) for e in listserv if e in removals]

    print(f"Sheet 1 - Cleaned Listserv : {len(kept)} addresses")
    print(f"Sheet 2 - Needs Review     : {len(review_rows)} addresses")
    print(f"Sheet 3 - Removed (bounce) : {len(removed_rows)} addresses\n")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Cleaned Listserv ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Cleaned Listserv"
    style_header(ws1, ["#", "Email Address"])
    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 45
    for i, email in enumerate(kept, 1):
        ws1.cell(row=i + 1, column=1, value=i)
        ws1.cell(row=i + 1, column=2, value=email)

    # ── Sheet 2: Needs Review ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Needs Review")
    style_header(ws2, ["#", "Email Address", "Sender Name",
                        "Category", "Summary (from Claude)", "Action Needed"])
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 25
    ws2.column_dimensions["D"].width = 28
    ws2.column_dimensions["E"].width = 60
    ws2.column_dimensions["F"].width = 30

    for i, (email, category, name, summary, action) in enumerate(review_rows, 1):
        r = i + 1
        for col_idx, val in enumerate(
                [i, email, name, category, summary, action], 1):
            cell = ws2.cell(row=r, column=col_idx, value=val)
            cell.fill = REVIEW_FILL
            cell.alignment = Alignment(wrap_text=True)
        ws2.row_dimensions[r].height = 45

    # ── Sheet 3: Removed (bounces) ────────────────────────────────────────────
    ws3 = wb.create_sheet("Removed (Bounces)")
    style_header(ws3, ["#", "Email Address", "Reason"])
    ws3.column_dimensions["B"].width = 45
    ws3.column_dimensions["C"].width = 30

    for i, (email, reason) in enumerate(removed_rows, 1):
        r = i + 1
        for col_idx, val in enumerate([i, email, reason], 1):
            cell = ws3.cell(row=r, column=col_idx, value=val)
            cell.fill = BOUNCE_FILL

    os.makedirs("output", exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"Saved -> {OUTPUT_FILE}")
    print(f"Generated : {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")


if __name__ == "__main__":
    main()
