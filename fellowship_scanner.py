"""
Science Corps Fellowship Scanner
=================================
Automated pipeline — Gmail fetch, bounce auto-tagging, No Reply cross-referencing,
and Excel export. Classification of remaining emails is done manually via Claude.ai.

Workflow
--------
1. Run four Gmail searches and merge results (inbox + spam).
2. Deduplicate by Gmail message ID.
3. Skip emails you sent yourself.
4. Auto-classify mailer-daemon / postmaster / -owner senders as Bounce.
5. Cross-reference listserv CSV to add No Reply rows.
6. Export auto-classified rows to fellowship_replies.xlsx (bounces + No Reply filled in).
7. Write pending emails to output/pending_for_claude.txt — paste into Claude.ai to classify.
8. Run fill_classifications.py to paste Claude's output back and complete the Excel report.
"""

import os
import re
import sys
import base64
import html
import datetime
from email.utils import parseaddr

# ─── CONFIG ──────────────────────────────────────────────────────────────────
SEARCH_KEYWORD          = "Paid Teaching Fellowship Abroad for Recent STEM PhDs"
SEARCH_AFTER_DATE       = "2026/06/01"           # Format: YYYY/MM/DD — update each cycle
MY_EMAIL                = "pnayga@science-corps.org"
# Addresses CC'd on the blast — excluded from NDR extraction
CC_EMAILS               = {
    "ccorry@science-corps.org",
    "cjellareroma@gmail.com",
    "ereroma@science-corps.org"
}

OUTPUT_FOLDER           = "output"
EXCEL_FILENAME          = "fellowship_replies.xlsx"
PENDING_FILENAME        = "pending_for_claude.xlsx"

LISTSERV_CSV            = "Filtered Listserv for Feb 2026 Applications.xlsx"
FROM_SEARCH_BATCH_SIZE  = 20     # listserv addresses per from: query batch
# ─────────────────────────────────────────────────────────────────────────────

# Gmail OAuth2 — read-only is enough
SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]

GMAIL_CATEGORY_LABELS = {
    "CATEGORY_PERSONAL": "Primary",
    "CATEGORY_UPDATES":  "Updates",
}

BOUNCE_SENDER_LOCALS = {"mailer-daemon", "postmaster"}

# Classification categories → Excel row fill colour (openpyxl hex, no #)
CATEGORY_COLORS = {
    "Positive / Interested":     "C6EFCE",  # green
    "Bounce / Delivery Failure": "FFC7CE",  # light red
    "No Longer Works There":     "FFEB9C",  # orange-yellow
    "Declined / Not Interested": "FF0000",  # dark red  (font goes white)
    "Has a Question":            "FFFF99",  # yellow
    "Application Submitted":     "BDD7EE",  # blue
    "Auto-Reply":                "E2CFFF",  # light purple
    "No Reply":                  "D9D9D9",  # light grey
    "Unverified":                "F2F2F2",  # very light grey
    "Pending":                   "FFFFE0",  # light yellow — awaiting Claude.ai classification
}
DARK_RED_FONT_WHITE = {"Declined / Not Interested"}

# Regex to parse one result line from Claude's output (shared with fill_classifications.py)
RESULT_LINE_RE = re.compile(
    r"Email\s+#?(\d+)\s*\|"
    r"\s*Category:\s*([^|]+?)\s*\|"
    r"\s*Summary:\s*([^|]+?)\s*\|"
    r"\s*Action:\s*(.+)",
    re.IGNORECASE,
)

# Category alias normalisation (shared with fill_classifications.py)
CATEGORY_ALIASES = {
    "positive":                       "Positive / Interested",
    "positive / interested":          "Positive / Interested",
    "interested":                     "Positive / Interested",
    "bounce":                         "Bounce / Delivery Failure",
    "bounce / delivery":              "Bounce / Delivery Failure",
    "bounce / delivery failure":      "Bounce / Delivery Failure",
    "delivery failure":               "Bounce / Delivery Failure",
    "no longer works":                "No Longer Works There",
    "no longer works there":          "No Longer Works There",
    "no longer there":                "No Longer Works There",
    "declined":                       "Declined / Not Interested",
    "declined / not interested":      "Declined / Not Interested",
    "not interested":                 "Declined / Not Interested",
    "has a question":                 "Has a Question",
    "question":                       "Has a Question",
    "application submitted":          "Application Submitted",
    "submitted":                      "Application Submitted",
    "auto-reply":                     "Auto-Reply",
    "auto reply":                     "Auto-Reply",
    "autoreply":                      "Auto-Reply",
    "out of office":                  "Auto-Reply",
    "out-of-office":                  "Auto-Reply",
    "automated reply":                "Auto-Reply",
    "automatic reply":                "Auto-Reply",
}


def normalise_category(raw):
    """Strip emojis / whitespace and map to a canonical category name."""
    cleaned = re.sub(r"[✅📧🚪❌❓📋]", "", raw).strip().lower()
    return CATEGORY_ALIASES.get(cleaned, raw.strip().title())


# ─── THIRD-PARTY IMPORTS ─────────────────────────────────────────────────────
try:
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request as GoogleRequest
    from googleapiclient.discovery import build
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"❌  Missing dependency: {e}")
    print("    Run:  pip install -r requirements.txt")
    sys.exit(1)


# ════════════════════════════════════════════════════════════════════════════
# PART 1 — GMAIL AUTHENTICATION & EMAIL FETCHING
# ════════════════════════════════════════════════════════════════════════════

def authenticate_gmail():
    creds = None
    token_path = "token.json"
    creds_path = "credentials.json"

    if not os.path.exists(creds_path):
        print("❌  credentials.json not found.")
        print("    Download it from Google Cloud Console → APIs & Services → Credentials.")
        sys.exit(1)

    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            print("🔄  Refreshing access token...")
            creds.refresh(GoogleRequest())
        else:
            print("🌐  Opening browser for Gmail authorisation...")
            flow = InstalledAppFlow.from_client_secrets_file(creds_path, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(token_path, "w") as f:
            f.write(creds.to_json())
        print("✅  Token saved to token.json")

    service = build("gmail", "v1", credentials=creds)
    print("✅  Authenticated with Gmail.\n")
    return service


def _paginate_search(service, query, include_spam_trash=False):
    ids = []
    next_page = None
    while True:
        kwargs = {"userId": "me", "q": query, "maxResults": 500,
                  "includeSpamTrash": include_spam_trash}
        if next_page:
            kwargs["pageToken"] = next_page
        try:
            resp = service.users().messages().list(**kwargs).execute()
        except Exception as e:
            print(f"   ⚠️  Gmail search error: {e}")
            break
        ids.extend(resp.get("messages", []))
        next_page = resp.get("nextPageToken")
        if not next_page:
            break
    return ids


def load_listserv(path):
    emails = set()
    if not os.path.exists(path):
        print(f"⚠️  Listserv file not found: {path}  — No Reply tracking disabled.")
        return emails
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                addr = str(cell or "").strip().lower()
                if "@" in addr:
                    emails.add(addr)
        wb.close()
    else:
        with open(path, "r", encoding="utf-8-sig") as f:
            for line in f:
                addr = line.strip().lower()
                if addr and "@" in addr:
                    emails.add(addr)
    print(f"📋  Loaded {len(emails)} addresses from listserv.")
    return emails


def get_sent_recipients(service):
    query = f'in:sent "{SEARCH_KEYWORD}" after:{SEARCH_AFTER_DATE}'
    print(f"📤  Checking Sent folder for original blast...")
    sent_ids = _paginate_search(service, query)
    print(f"    → {len(sent_ids)} sent email(s) found")

    recipients = set()
    for msg_ref in sent_ids:
        try:
            msg = service.users().messages().get(
                userId="me", id=msg_ref["id"], format="metadata",
                metadataHeaders=["To", "Cc", "Bcc"]
            ).execute()
            headers = {h["name"].lower(): h["value"]
                       for h in msg.get("payload", {}).get("headers", [])}
            for field in ["to", "cc", "bcc"]:
                raw = headers.get(field, "")
                if not raw:
                    continue
                for part in raw.split(","):
                    _, addr = parseaddr(part.strip())
                    if addr:
                        recipients.add(addr.lower().strip())
        except Exception as e:
            print(f"   ⚠️  Could not read sent message: {e}")

    if recipients:
        print(f"    → {len(recipients)} unique recipient(s) found in sent records.\n")
    else:
        print(f"    ⚠️  No recipients found — BCC records may not be accessible.\n")
    return recipients


def _search_from_listserv(service, listserv_emails, include_spam_trash=False):
    email_list = sorted(listserv_emails)
    batches = [email_list[i:i + FROM_SEARCH_BATCH_SIZE]
               for i in range(0, len(email_list), FROM_SEARCH_BATCH_SIZE)]
    all_ids = []
    for idx, batch in enumerate(batches, 1):
        from_clause = " OR ".join(f"from:{e}" for e in batch)
        query = f"({from_clause}) after:{SEARCH_AFTER_DATE}"
        ids = _paginate_search(service, query, include_spam_trash=include_spam_trash)
        all_ids.extend(ids)
        if idx % 10 == 0 or idx == len(batches):
            print(f"   ... {idx}/{len(batches)} batches done ({len(all_ids)} found so far)")
    return all_ids


def get_gmail_category(label_ids):
    for label_id, name in GMAIL_CATEGORY_LABELS.items():
        if label_id in label_ids:
            return name
    return "Other"


def decode_body(part):
    data = part.get("body", {}).get("data", "")
    if not data:
        return ""
    try:
        return base64.urlsafe_b64decode(data + "==").decode("utf-8", errors="replace")
    except Exception:
        return ""


def strip_html(raw_html):
    clean = re.sub(r"<(style|script)[^>]*>.*?</(style|script)>", "",
                   raw_html, flags=re.DOTALL | re.IGNORECASE)
    clean = re.sub(r"<[^>]+>", " ", clean)
    clean = re.sub(r"\s+", " ", clean).strip()
    return html.unescape(clean)


def extract_body(payload):
    mime = payload.get("mimeType", "")
    if mime == "text/plain":
        return decode_body(payload)
    if mime == "text/html":
        return strip_html(decode_body(payload))
    parts = payload.get("parts", [])
    plain, html_text = "", ""
    for part in parts:
        result = extract_body(part)
        if part.get("mimeType") == "text/plain" and result:
            plain = result
        elif part.get("mimeType") in ("text/html", "multipart/alternative") and result:
            html_text = result
        elif not plain and result:
            plain = result
    return plain or html_text


def auto_classify_sender(sender_email):
    """
    Returns {category, summary, action} if sender can be auto-classified,
    or None if manual classification via Claude.ai is needed.
    """
    local = sender_email.split("@")[0].lower().strip()
    if local in BOUNCE_SENDER_LOCALS:
        return {
            "category": "Bounce / Delivery Failure",
            "summary":  "Automated delivery failure / NDR from mail server.",
            "action":   "Remove from list",
        }
    if local.endswith("-owner"):
        return {
            "category": "Bounce / Delivery Failure",
            "summary":  "Mailing list owner rejection or moderation notice.",
            "action":   "Remove from list",
        }
    return None


def fetch_emails(service, listserv=None):
    query_main        = f'"{SEARCH_KEYWORD}" after:{SEARCH_AFTER_DATE} -from:{MY_EMAIL}'
    query_bounce      = f'(from:mailer-daemon OR from:postmaster) after:{SEARCH_AFTER_DATE}'
    query_bounce_subj = (
        f'(subject:undeliverable OR subject:"delivery failed" OR '
        f'subject:"address not found" OR subject:"delivery status notification" OR '
        f'subject:"returned mail" OR subject:"mail delivery failure" OR '
        f'subject:"failure notice" OR subject:"undelivered mail" OR '
        f'subject:"mail delivery subsystem") '
        f'after:{SEARCH_AFTER_DATE} -from:{MY_EMAIL}'
    )

    print(f"🔍  Main search (inbox + spam):\n    {query_main}")
    main_ids      = _paginate_search(service, query_main)
    main_ids_spam = _paginate_search(service, query_main, include_spam_trash=True)
    print(f"    → {len(main_ids)} inbox / {len(main_ids_spam)} incl. spam\n")

    print(f"🔍  Bounce sender search (inbox + spam):\n    {query_bounce}")
    bounce_ids      = _paginate_search(service, query_bounce)
    bounce_ids_spam = _paginate_search(service, query_bounce, include_spam_trash=True)
    print(f"    → {len(bounce_ids)} inbox / {len(bounce_ids_spam)} incl. spam\n")

    print(f"🔍  Bounce subject search (inbox + spam):\n    {query_bounce_subj}")
    bounce_subj_ids      = _paginate_search(service, query_bounce_subj)
    bounce_subj_ids_spam = _paginate_search(service, query_bounce_subj, include_spam_trash=True)
    print(f"    → {len(bounce_subj_ids)} inbox / {len(bounce_subj_ids_spam)} incl. spam\n")

    listserv_ids = []
    if listserv:
        print(f"🔍  Listserv from-search: {len(listserv)} addresses "
              f"→ {len(listserv) // FROM_SEARCH_BATCH_SIZE + 1} batch queries (inbox)...")
        listserv_ids_inbox = _search_from_listserv(service, listserv, include_spam_trash=False)
        print(f"    → {len(listserv_ids_inbox)} inbox\n"
              f"🔍  Listserv from-search (incl. spam)...")
        listserv_ids_spam  = _search_from_listserv(service, listserv, include_spam_trash=True)
        print(f"    → {len(listserv_ids_spam)} incl. spam\n")
        listserv_ids = listserv_ids_inbox + listserv_ids_spam

    seen, combined = set(), []
    for ref in (main_ids + main_ids_spam + bounce_ids + bounce_ids_spam +
                bounce_subj_ids + bounce_subj_ids_spam + listserv_ids):
        if ref["id"] not in seen:
            seen.add(ref["id"])
            combined.append(ref)

    total_unique = len(combined)
    if total_unique == 0:
        print("📭  No emails found. Check SEARCH_KEYWORD and SEARCH_AFTER_DATE.")
        sys.exit(0)

    print(f"📬  {total_unique} unique email(s) after deduplication. Fetching content...\n")

    emails, skipped = [], 0

    for msg_ref in combined:
        try:
            msg = service.users().messages().get(
                userId="me", id=msg_ref["id"], format="full"
            ).execute()
        except Exception as e:
            print(f"   ⚠️  Skipping {msg_ref['id']}: {e}")
            skipped += 1
            continue

        headers = {h["name"].lower(): h["value"]
                   for h in msg.get("payload", {}).get("headers", [])}

        from_raw = headers.get("from", "")
        sender_name, sender_email = parseaddr(from_raw)
        sender_email = sender_email.lower().strip()

        if sender_email == MY_EMAIL.lower().strip():
            skipped += 1
            continue

        email_num = len(emails) + 1
        emails.append({
            "number":         email_num,
            "sender_name":    sender_name or sender_email,
            "sender_email":   sender_email,
            "date":           headers.get("date", ""),
            "subject":        headers.get("subject", "(no subject)"),
            "body":           extract_body(msg.get("payload", {})) or "(body unavailable)",
            "gmail_category": get_gmail_category(msg.get("labelIds", [])),
        })
        print(f"   ✅  Email {email_num}/{total_unique - skipped} — {sender_email}")

    if skipped:
        print(f"\n   ℹ️  Skipped {skipped} email(s) (outgoing or fetch error).")

    for i, e in enumerate(emails, start=1):
        e["number"] = i

    return emails


# ════════════════════════════════════════════════════════════════════════════
# PART 2 — AUTO-CLASSIFICATION & PENDING FILE
# ════════════════════════════════════════════════════════════════════════════

def save_pending_xlsx(pending_emails, output_path):
    """
    Export unclassified emails to a plain Excel file for upload to Claude.ai.
    Columns: #, Sender Name, Email Address, Institution, Date Received,
             Gmail Category, Subject, Body Preview (first 800 chars).
    """
    if not pending_emails:
        print("   ℹ️  No pending emails — skipping pending_for_claude.xlsx.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pending Emails"

    HEADERS = ["#", "Sender Name", "Email Address", "Institution",
               "Date Received", "Gmail Category", "Subject", "Body Preview"]
    ws.append(HEADERS)

    hfill = PatternFill(fill_type="solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    for e in pending_emails:
        body_preview = e["body"][:800]
        if len(e["body"]) > 800:
            body_preview += "... [truncated]"
        ws.append([
            e["number"],
            e["sender_name"],
            e["sender_email"],
            extract_institution(e["sender_email"], e["sender_name"], e["body"]),
            e["date"],
            e["gmail_category"],
            e["subject"],
            body_preview,
        ])
        row_idx = ws.max_row
        ws.cell(row=row_idx, column=8).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_idx].height = 80

    for i, w in enumerate([5, 25, 35, 25, 28, 14, 40, 80], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"   💾  Pending emails saved → {output_path}")


def classify_emails(emails):
    """
    Auto-classify what we can; mark the rest as Pending.
    Returns (results_list, pending_emails_list).
    """
    results = {}
    pending = []

    for e in emails:
        auto = auto_classify_sender(e["sender_email"])
        if auto:
            results[e["number"]] = {
                "number":         e["number"],
                "sender_name":    e["sender_name"],
                "sender_email":   e["sender_email"],
                "date":           e["date"],
                "subject":        e["subject"],
                "gmail_category": e["gmail_category"],
                "category":       auto["category"],
                "summary":        auto["summary"],
                "action":         auto["action"],
                "institution":    extract_institution(e["sender_email"], e["sender_name"], e["body"]),
            }
        else:
            results[e["number"]] = {
                "number":         e["number"],
                "sender_name":    e["sender_name"],
                "sender_email":   e["sender_email"],
                "date":           e["date"],
                "subject":        e["subject"],
                "gmail_category": e["gmail_category"],
                "category":       "Pending",
                "summary":        "",
                "action":         "",
                "institution":    extract_institution(e["sender_email"], e["sender_name"], e["body"]),
            }
            pending.append(e)

    auto_count = len(emails) - len(pending)
    print(f"\n   ✅  Auto-tagged {auto_count} email(s) (bounces / list rejections).")
    print(f"   📝  {len(pending)} email(s) marked Pending — will go to pending_for_claude.xlsx.")

    return [results[k] for k in sorted(results)], pending


# ════════════════════════════════════════════════════════════════════════════
# HELPERS
# ════════════════════════════════════════════════════════════════════════════

FREE_DOMAINS = {
    "gmail.com", "yahoo.com", "hotmail.com", "outlook.com",
    "aol.com", "icloud.com", "protonmail.com", "mail.com",
}


def extract_institution(sender_email, sender_name, body):
    domain = sender_email.split("@")[-1].lower() if "@" in sender_email else ""
    if domain and domain not in FREE_DOMAINS:
        return domain.replace(".edu", "").replace(".org", "").replace("-", " ").title()
    snippet = body[:500]
    for pat in [
        r"(University of [A-Z][a-z]+(?: [A-Z][a-z]+)*)",
        r"([A-Z][a-z]+(?: [A-Z][a-z]+)* University)",
        r"([A-Z][a-z]+(?: [A-Z][a-z]+)* College)",
        r"([A-Z][a-z]+(?: [A-Z][a-z]+)* Institute)",
    ]:
        m = re.search(pat, snippet)
        if m:
            return m.group(1)
    return ""


_EMAIL_RE = re.compile(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}')


def extract_failed_recipients(body):
    return {addr.lower() for addr in _EMAIL_RE.findall(body)}


def _thin_border():
    thin = Side(style="thin", color="AAAAAA")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ════════════════════════════════════════════════════════════════════════════
# PART 3 — EXCEL EXPORT
# ════════════════════════════════════════════════════════════════════════════

def build_action_plan_sheet(wb, ws_replies):
    STEP1_CATS = {"Positive / Interested", "Has a Question", "Application Submitted"}
    STEP2_CATS = {"No Longer Works There", "Declined / Not Interested"}
    STEP3_CATS = {"Auto-Reply", "No Reply", "Unverified"}

    step1, step2 = [], []
    step3_count = step4_count = 0

    for row in ws_replies.iter_rows(min_row=2, values_only=True):
        if not isinstance(row[0], (int, float)):
            continue
        category = str(row[6] or "")
        entry = (str(row[1] or ""), str(row[2] or ""), str(row[3] or ""),
                 str(row[4] or ""), category, str(row[7] or ""), str(row[8] or ""))
        if category in STEP1_CATS:
            step1.append(entry)
        elif category in STEP2_CATS:
            step2.append(entry)
        elif category in STEP3_CATS:
            step3_count += 1
        else:
            step4_count += 1

    if "Action Plan" in wb.sheetnames:
        del wb["Action Plan"]
    ws = wb.create_sheet("Action Plan")

    DETAIL_COLS = ["Sender Name", "Email Address", "Institution",
                   "Date Received", "Category", "Claude's Summary", "Recommended Action"]

    def section_hdr(rn, step_num, label, count, bg):
        title = f"STEP {step_num} — {label}  ({count} email{'s' if count != 1 else ''})"
        c = ws.cell(row=rn, column=1, value=title)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(bold=True, size=12, color="FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"A{rn}:G{rn}")
        ws.row_dimensions[rn].height = 22
        return rn + 1

    def col_hdr(rn):
        for ci, h in enumerate(DETAIL_COLS, 1):
            c = ws.cell(row=rn, column=ci, value=h)
            c.fill = PatternFill("solid", fgColor="2E75B6")
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")
            c.border = _thin_border()
        ws.row_dimensions[rn].height = 18
        return rn + 1

    def data_rows(rn, rows):
        if not rows:
            c = ws.cell(row=rn, column=1, value="  No emails in this category.")
            c.font = Font(italic=True, color="888888")
            ws.merge_cells(f"A{rn}:G{rn}")
            return rn + 1
        for (sname, email, inst, date, cat, summary, action) in rows:
            fhex = CATEGORY_COLORS.get(cat, "FFFFFF")
            for ci, val in enumerate([sname, email, inst, date, cat, summary, action], 1):
                c = ws.cell(row=rn, column=ci, value=val)
                c.fill = PatternFill("solid", fgColor=fhex)
                c.border = _thin_border()
                c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[rn].height = 55
            rn += 1
        return rn

    def note_row(rn, text):
        c = ws.cell(row=rn, column=1, value=text)
        c.font = Font(italic=True, color="444444")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"A{rn}:G{rn}")
        ws.row_dimensions[rn].height = 38
        return rn + 2

    ws["A1"] = "Science Corps Fellowship — Action Plan"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws.merge_cells("A1:G1")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"] = (f"Generated: {datetime.date.today().strftime('%B %d, %Y')}  |  "
                "Work through each step in order.")
    ws["A2"].font = Font(italic=True, color="595959")
    ws.merge_cells("A2:G2")

    rn = 4
    rn = section_hdr(rn, 1, "FOLLOW UP NOW", len(step1), "1A7A3C")
    rn = col_hdr(rn)
    rn = data_rows(rn, step1)
    rn += 1

    rn = section_hdr(rn, 2, "SPECIAL ACTION NEEDED", len(step2), "BF6A00")
    rn = col_hdr(rn)
    rn = data_rows(rn, step2)
    rn += 1

    rn = section_hdr(rn, 3, "INCLUDE IN NEXT BLAST", step3_count, "595959")
    rn = note_row(rn,
        f"  {step3_count} addresses did not reply or were out of office. "
        "Run clean_listserv.py to generate the filtered blast list for the next cycle.")

    rn = section_hdr(rn, 4, "DEAD ADDRESSES — REMOVE FROM LIST", step4_count, "9C0006")
    note_row(rn,
        f"  {step4_count} hard bounces — these addresses cannot receive mail. "
        "Run clean_listserv.py to strip them from your listserv before the next blast.")

    for ci, w in enumerate([25, 38, 25, 22, 24, 60, 50], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A4"


def load_existing_excel(path):
    if not os.path.exists(path):
        return None, set()
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        print(f"   ⚠️  Could not open existing file ({e}). Starting fresh.")
        return None, set()
    existing = set()
    if "Replies" in wb.sheetnames:
        for row in wb["Replies"].iter_rows(min_row=2, values_only=True):
            if row[2]:
                existing.add((str(row[2]).lower().strip(), str(row[4]).strip() if row[4] else ""))
    return wb, existing


def build_summary_sheet(wb, ws_replies):
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws_s = wb.create_sheet("Summary", 0)

    cat_counts, gcat_counts, total_rows = {}, {}, 0
    for row in ws_replies.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        total_rows += 1
        cat_counts[str(row[6] or "Unclassified")] = cat_counts.get(str(row[6] or "Unclassified"), 0) + 1
        gcat_counts[str(row[5] or "Other")]        = gcat_counts.get(str(row[5] or "Other"), 0) + 1

    ws_s.column_dimensions["A"].width = 32
    ws_s.column_dimensions["B"].width = 12

    title_font   = Font(bold=True, size=14, color="1F4E79")
    section_fill = PatternFill(fill_type="solid", fgColor="2E75B6")
    section_font = Font(bold=True, size=11, color="FFFFFF")

    ws_s["A1"] = "Science Corps Fellowship Scanner — Report"
    ws_s["A1"].font = title_font
    ws_s.merge_cells("A1:B1")
    ws_s["A1"].alignment = Alignment(horizontal="center")
    ws_s["A2"] = "Generated:"
    ws_s["B2"] = datetime.date.today().strftime("%B %d, %Y")
    ws_s["A3"] = "Total emails processed:"
    ws_s["B3"] = total_rows
    ws_s["A3"].font = Font(bold=True)
    ws_s.append([])

    ws_s.append(["Classification Breakdown", "Count"])
    for cell in ws_s[ws_s.max_row]:
        cell.fill = section_fill; cell.font = section_font
        cell.alignment = Alignment(horizontal="center"); cell.border = _thin_border()
    for cat, cnt in sorted(cat_counts.items()):
        ws_s.append([cat, cnt])
        for cell in ws_s[ws_s.max_row]:
            cell.fill   = PatternFill(fill_type="solid", fgColor=CATEGORY_COLORS.get(cat, "FFFFFF"))
            cell.border = _thin_border()
            if cat in DARK_RED_FONT_WHITE:
                cell.font = Font(color="FFFFFF")

    ws_s.append([])
    ws_s.append(["Gmail Category Breakdown", "Count"])
    for cell in ws_s[ws_s.max_row]:
        cell.fill = section_fill; cell.font = section_font
        cell.alignment = Alignment(horizontal="center"); cell.border = _thin_border()
    for gcat, cnt in sorted(gcat_counts.items()):
        ws_s.append([gcat, cnt])
        for cell in ws_s[ws_s.max_row]:
            cell.border = _thin_border()

    return total_rows, cat_counts, gcat_counts


def export_to_excel(results, output_path):
    print(f"\n📊  Exporting to Excel...")

    wb, existing = load_existing_excel(output_path)
    if wb is None:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    HEADERS = ["#", "Sender Name", "Email Address", "Institution",
               "Date Received", "Gmail Category", "Category", "Summary", "Action Needed"]

    if "Replies" not in wb.sheetnames:
        ws = wb.create_sheet("Replies")
    else:
        ws = wb["Replies"]

    if ws.max_row == 1 and not any(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        ws.append(HEADERS)
        hfill = PatternFill(fill_type="solid", fgColor="1F4E79")
        hfont = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[1]:
            cell.fill = hfill; cell.font = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _thin_border()
        ws.row_dimensions[1].height = 20

    added = skipped_dupes = 0

    for r in results:
        key = (r["sender_email"].lower().strip(), r["date"].strip())
        if key in existing:
            skipped_dupes += 1
            continue

        ws.append([
            r["number"], r["sender_name"], r["sender_email"], r["institution"],
            r["date"], r["gmail_category"], r["category"], r["summary"], r["action"],
        ])
        added += 1
        existing.add(key)

        fill_hex = CATEGORY_COLORS.get(r["category"], "FFFFFF")
        row_idx  = ws.max_row
        for col in range(1, len(HEADERS) + 1):
            cell           = ws.cell(row=row_idx, column=col)
            cell.fill      = PatternFill(fill_type="solid", fgColor=fill_hex)
            cell.border    = _thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if r["category"] in DARK_RED_FONT_WHITE:
                cell.font  = Font(color="FFFFFF")

    for i, w in enumerate([5, 25, 35, 25, 28, 14, 26, 55, 30], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    total_rows, cat_counts, gcat_counts = build_summary_sheet(wb, ws)
    build_action_plan_sheet(wb, ws)
    wb._sheets.insert(0, wb._sheets.pop(wb.sheetnames.index("Action Plan")))

    wb.save(output_path)
    print(f"✅  Report saved → {output_path}")
    if skipped_dupes:
        print(f"   ℹ️  {skipped_dupes} duplicate(s) skipped.")
    print(f"   ℹ️  {added} new row(s) added.")
    return total_rows, cat_counts, gcat_counts


# ════════════════════════════════════════════════════════════════════════════
# PART 4 — TERMINAL SUMMARY
# ════════════════════════════════════════════════════════════════════════════

def print_final_summary(emails, total_rows, cat_counts, excel_path, pending_path, pending_count):
    print("\n" + "=" * 62)
    print("  FINAL SUMMARY")
    print("=" * 62)
    print(f"  Emails fetched this run  : {len(emails)}")

    print("\n  Classification (full report):")
    for k, v in sorted(cat_counts.items()):
        print(f"    {k:<32} {v}")

    print(f"\n  Total rows in Excel       : {total_rows}")
    print(f"  Report                    : {excel_path}")
    print("=" * 62)

    if pending_count > 0:
        print(f"""
┌─────────────────────────────────────────────────────────────┐
│  NEXT STEP — CLASSIFY IN CLAUDE.AI                          │
├─────────────────────────────────────────────────────────────┤
│                                                             │
│  {pending_count} email(s) need classification.                          │
│                                                             │
│  1. Open the Science Corps Classification project           │
│     in Claude.ai                                            │
│  2. Upload:  output\\pending_for_claude.xlsx                │
│  3. Claude will classify each email and return a            │
│     completed fellowship_replies.xlsx for download          │
│  4. Save the downloaded file to:  output\\                  │
│     fellowship_replies.xlsx                                 │
│                                                             │
└─────────────────────────────────────────────────────────────┘
""")
    else:
        print("\n  ✅  All emails auto-classified. No pending file generated.\n")


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════

def main():
    print("\n" + "═" * 62)
    print("  Science Corps Fellowship Scanner")
    print("═" * 62 + "\n")

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    listserv = load_listserv(LISTSERV_CSV) if LISTSERV_CSV else set()

    service = authenticate_gmail()
    sent_recipients = get_sent_recipients(service)
    emails = fetch_emails(service, listserv)

    print(f"\n🏷️   Auto-classifying {len(emails)} email(s)...\n")
    results, pending_emails = classify_emails(emails)

    # ── Extract failed delivery addresses from bounce NDR bodies ─────────────
    email_body_map   = {e["number"]: e.get("body", "") for e in emails}
    bounced_listserv = set()
    if listserv:
        for r in results:
            if r.get("category") == "Bounce / Delivery Failure":
                body     = email_body_map.get(r["number"], "")
                excluded = {MY_EMAIL.lower()} | {e.lower() for e in CC_EMAILS}
                for addr in extract_failed_recipients(body):
                    if addr in listserv and addr not in excluded:
                        bounced_listserv.add(addr)
        if bounced_listserv:
            print(f"   ✅  {len(bounced_listserv)} listserv address(es) confirmed bounced "
                  f"(extracted from NDR bodies).")

    # ── Add No Reply rows for listserv addresses that didn't reply ────────────
    if listserv:
        replied          = {r["sender_email"].lower().strip() for r in results}
        no_reply_count   = bounce_confirmed = unverified_count = 0

        for addr in sorted(listserv):
            if addr not in replied:
                if addr in bounced_listserv:
                    category = "Bounce / Delivery Failure"
                    summary  = "Bounced — address could not receive mail (confirmed via NDR body)."
                    action   = "Remove from list"
                    bounce_confirmed += 1
                elif sent_recipients:
                    if addr in sent_recipients:
                        category = "No Reply"
                        summary  = "Email confirmed sent. No reply received."
                        action   = "Follow up if needed"
                    else:
                        category = "Unverified"
                        summary  = "Address not found in sent records — email may not have been delivered."
                        action   = "Verify address and re-send if needed"
                        unverified_count += 1
                else:
                    category = "No Reply"
                    summary  = "No reply received. Could not verify delivery (BCC records unavailable)."
                    action   = "None"

                results.append({
                    "number": 0, "sender_name": "", "sender_email": addr,
                    "institution": extract_institution(addr, "", ""),
                    "date": "", "gmail_category": "",
                    "category": category, "summary": summary, "action": action,
                })
                no_reply_count += 1

        print(f"   ℹ️  Listserv: {bounce_confirmed} confirmed bounced, "
              f"{no_reply_count - bounce_confirmed - unverified_count} no reply, "
              f"{unverified_count} unverified.")

        for i, r in enumerate(results, 1):
            r["number"] = i

    # ── Save pending xlsx ─────────────────────────────────────────────────────
    pending_path = os.path.join(OUTPUT_FOLDER, PENDING_FILENAME)
    save_pending_xlsx(pending_emails, pending_path)

    # ── Export Excel ──────────────────────────────────────────────────────────
    print(f"\n📋  Total rows to write: {len(results)}")
    excel_path = os.path.join(OUTPUT_FOLDER, EXCEL_FILENAME)
    total_rows, cat_counts, gcat_counts = export_to_excel(results, excel_path)

    print_final_summary(emails, total_rows, cat_counts, excel_path,
                        pending_path, len(pending_emails))


if __name__ == "__main__":
    main()
