"""
positive_replies_report.py
--------------------------
Reads fellowship_replies.xlsx and generates an HTML report for every
Positive / Interested reply. Each entry shows:
  - Sender name, email, institution, date
  - Claude's summary of their reply
  - A ready-to-send draft thank-you + network inquiry email

Open the output HTML in Chrome/Edge and press Ctrl+P -> Save as PDF.
"""

import os
import sys
import datetime
import html as htmllib

try:
    import openpyxl
except ImportError:
    print("Missing dependency: openpyxl  ->  pip install openpyxl")
    sys.exit(1)

# ─── CONFIG ──────────────────────────────────────────────────────────────────
EXCEL_INPUT  = os.path.join("output", "fellowship_replies.xlsx")
OUTPUT_HTML  = os.path.join("output", "positive_replies_outreach.html")
YOUR_NAME    = "Prince Nayga"
YOUR_TITLE   = "Philippine Manager, Science Corps"
YOUR_EMAIL   = "pnayga@science-corps.org"
ORG_NAME     = "Science Corps"
FELLOWSHIP   = "Paid Teaching Fellowship Abroad for Recent STEM PhDs"
# ─────────────────────────────────────────────────────────────────────────────

TARGET_CATEGORY = "Positive / Interested"

TEMPLATE = """\
Dear {first_name},

Thank you so much for sharing the Science Corps Teaching Fellowship opportunity with your
students! We truly appreciate your support — it makes a real difference in helping us reach
talented STEM PhD graduates who could benefit from this experience.

As we continue to grow our outreach, we are looking to expand our network of institutional
contacts. Would you happen to know of any other departmental or program mailing list addresses
— similar to yours — that we could reach out to directly? For example, graduate program
coordinators, department listservs, or student association emails at your institution or
others you are connected with.

If you could share any such email addresses, we would be very grateful. We will make sure to
reach out to them individually and respectfully, just as we did with you.

For reference, the fellowship we are promoting is:

  {fellowship}

Please feel free to reply to this email with any contacts or leads you may have. And of
course, if you have any questions or need more materials to share with your students, just
let me know — I am happy to help.

Thank you again for your generosity and support!

Warm regards,

{your_name}
{your_title}
{org_name}
{your_email}
"""


def load_positive_replies(excel_path):
    if not os.path.exists(excel_path):
        print(f"Excel report not found: {excel_path}")
        print("Run fellowship_scanner.py first.")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path)
    if "Replies" not in wb.sheetnames:
        print("No 'Replies' sheet found.")
        sys.exit(1)

    ws = wb["Replies"]
    positives = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not isinstance(row[0], (int, float)):
            continue
        category = str(row[6] or "")
        if category != TARGET_CATEGORY:
            continue
        positives.append({
            "name":        str(row[1] or "").strip(),
            "email":       str(row[2] or "").strip(),
            "institution": str(row[3] or "").strip(),
            "date":        str(row[4] or "").strip(),
            "summary":     str(row[7] or "").strip(),
            "action":      str(row[8] or "").strip(),
        })

    return positives


def first_name(full_name):
    """Extract first name, or fall back to 'there' if name is blank/email-like."""
    if not full_name or "@" in full_name:
        return "there"
    parts = full_name.strip().split()
    return parts[0].title() if parts else "there"


def build_draft(person):
    return TEMPLATE.format(
        first_name   = first_name(person["name"]),
        fellowship   = FELLOWSHIP,
        your_name    = YOUR_NAME,
        your_title   = YOUR_TITLE,
        your_email   = YOUR_EMAIL,
        org_name     = ORG_NAME,
    )


def esc(text):
    return htmllib.escape(str(text))


def build_html(positives):
    generated = datetime.datetime.now().strftime("%B %d, %Y at %I:%M %p")

    cards = ""
    for i, p in enumerate(positives, 1):
        draft = build_draft(p)
        cards += f"""
        <div class="card">
          <div class="card-header">
            <span class="badge">{i}</span>
            <div class="person-info">
              <div class="person-name">{esc(p['name'] or p['email'])}</div>
              <div class="person-meta">
                <span class="meta-item email-chip">{esc(p['email'])}</span>
                {"<span class='meta-item'>" + esc(p['institution']) + "</span>" if p['institution'] else ""}
                {"<span class='meta-item date'>" + esc(p['date'][:16]) + "</span>" if p['date'] else ""}
              </div>
            </div>
          </div>

          <div class="section-label">Their Reply (Claude's Summary)</div>
          <div class="summary-box">{esc(p['summary'])}</div>

          {"<div class='section-label'>Recommended Action</div><div class='action-box'>" + esc(p['action']) + "</div>" if p['action'] else ""}

          <div class="section-label">Draft Email — ready to send</div>
          <div class="draft-meta">
            <strong>To:</strong> {esc(p['email'])}<br>
            <strong>Subject:</strong> Thank you for sharing — do you know other contacts we can reach?
          </div>
          <div class="draft-box"><pre>{esc(draft)}</pre></div>
        </div>
        """

    if not cards:
        cards = "<p class='empty'>No Positive / Interested replies found in the report.</p>"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <title>Positive Replies — {esc(ORG_NAME)} Fellowship Outreach</title>
  <style>
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{
      font-family: 'Segoe UI', Arial, sans-serif;
      font-size: 13px;
      color: #222;
      background: #f4f6fb;
      padding: 32px 24px;
    }}
    h1 {{
      font-size: 22px;
      color: #1F4E79;
      margin-bottom: 4px;
    }}
    .subtitle {{
      color: #555;
      margin-bottom: 6px;
      font-size: 12px;
    }}
    .count-badge {{
      display: inline-block;
      background: #1A7A3C;
      color: #fff;
      border-radius: 12px;
      padding: 3px 14px;
      font-size: 12px;
      font-weight: bold;
      margin-bottom: 24px;
    }}
    .card {{
      background: #fff;
      border-radius: 8px;
      box-shadow: 0 2px 8px rgba(0,0,0,0.09);
      margin-bottom: 32px;
      overflow: hidden;
      page-break-inside: avoid;
    }}
    .card-header {{
      background: #1F4E79;
      color: #fff;
      padding: 14px 20px;
      display: flex;
      align-items: flex-start;
      gap: 14px;
    }}
    .badge {{
      background: #fff;
      color: #1F4E79;
      font-weight: bold;
      border-radius: 50%;
      width: 28px;
      height: 28px;
      display: flex;
      align-items: center;
      justify-content: center;
      font-size: 13px;
      flex-shrink: 0;
      margin-top: 2px;
    }}
    .person-name {{
      font-size: 16px;
      font-weight: bold;
      margin-bottom: 5px;
    }}
    .person-meta {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      align-items: center;
    }}
    .meta-item {{
      font-size: 11.5px;
      opacity: 0.9;
    }}
    .email-chip {{
      background: rgba(255,255,255,0.18);
      border-radius: 10px;
      padding: 2px 9px;
      font-family: monospace;
    }}
    .date {{ opacity: 0.7; font-style: italic; }}
    .section-label {{
      font-size: 10px;
      font-weight: bold;
      text-transform: uppercase;
      letter-spacing: 0.8px;
      color: #555;
      padding: 12px 20px 4px;
    }}
    .summary-box {{
      background: #f0f7f2;
      border-left: 4px solid #1A7A3C;
      margin: 0 20px 12px;
      padding: 10px 14px;
      border-radius: 0 4px 4px 0;
      line-height: 1.5;
      color: #1a4a2a;
    }}
    .action-box {{
      background: #fff8e1;
      border-left: 4px solid #BF6A00;
      margin: 0 20px 12px;
      padding: 10px 14px;
      border-radius: 0 4px 4px 0;
      line-height: 1.5;
      color: #5a3a00;
    }}
    .draft-meta {{
      margin: 4px 20px 6px;
      font-size: 12px;
      color: #333;
      line-height: 1.8;
    }}
    .draft-box {{
      background: #f7f9fc;
      border: 1px solid #dde3ee;
      border-radius: 6px;
      margin: 0 20px 20px;
      padding: 14px 16px;
    }}
    .draft-box pre {{
      white-space: pre-wrap;
      word-break: break-word;
      font-family: 'Segoe UI', Arial, sans-serif;
      font-size: 13px;
      line-height: 1.65;
      color: #222;
    }}
    .empty {{ color: #888; font-style: italic; padding: 20px; }}
    @media print {{
      body {{ background: white; padding: 0; }}
      .card {{ box-shadow: none; border: 1px solid #ccc; margin-bottom: 24px; }}
    }}
  </style>
</head>
<body>
  <h1>{esc(ORG_NAME)} Fellowship — Positive Replies &amp; Outreach Drafts</h1>
  <div class="subtitle">Generated: {generated}</div>
  <div class="count-badge">{len(positives)} Positive / Interested</div>

  {cards}
</body>
</html>
"""


def main():
    print(f"Reading : {EXCEL_INPUT}")
    positives = load_positive_replies(EXCEL_INPUT)
    print(f"Found   : {len(positives)} Positive / Interested reply/replies\n")

    os.makedirs("output", exist_ok=True)
    html_content = build_html(positives)
    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(html_content)

    print(f"Saved   : {OUTPUT_HTML}")
    print()
    print("Next steps:")
    print("  1. Open the HTML file in Chrome or Edge")
    print("  2. Press Ctrl+P  ->  'Save as PDF' to export")
    print("  3. Review and personalise each draft before sending")
    print()
    for i, p in enumerate(positives, 1):
        print(f"  {i}. {p['name'] or p['email']}  <{p['email']}>")


if __name__ == "__main__":
    main()
